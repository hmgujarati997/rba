from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import io
import csv
import uuid
import json
import base64
import logging
import secrets
import asyncio
from typing import List, Optional, Literal
from datetime import datetime, timezone, timedelta

import bcrypt
import jwt
import qrcode
import httpx
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, ConfigDict, EmailStr

# ---------- Setup ----------
JWT_ALGORITHM = "HS256"
JWT_EXPIRES_MIN = 60 * 24 * 7  # 7 days for ease of mobile PWA use

UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="Rama Bazaar 1.0 API")
api = APIRouter(prefix="/api")

app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")
app.mount("/api/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads_api")

logger = logging.getLogger("rama")
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# ---------- Helpers ----------
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False

def jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

def create_token(sub: str, role: str, extra: Optional[dict] = None) -> str:
    payload = {
        "sub": sub,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=JWT_EXPIRES_MIN),
        "iat": datetime.now(timezone.utc),
    }
    if extra:
        payload.update(extra)
    return jwt.encode(payload, jwt_secret(), algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    return jwt.decode(token, jwt_secret(), algorithms=[JWT_ALGORITHM])

def normalize_mobile(m: str) -> str:
    return "".join(ch for ch in (m or "") if ch.isdigit())[-10:]

async def get_current(request: Request) -> dict:
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = decode_token(auth[7:])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(user: dict = Depends(get_current)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return user

async def require_exhibitor(user: dict = Depends(get_current)) -> dict:
    if user.get("role") != "exhibitor":
        raise HTTPException(status_code=403, detail="Exhibitor only")
    return user

async def require_gate_or_admin(user: dict = Depends(get_current)) -> dict:
    if user.get("role") not in ("admin", "gate"):
        raise HTTPException(status_code=403, detail="Gate or admin only")
    return user

# ---------- Models ----------
class LoginIn(BaseModel):
    email: Optional[EmailStr] = None
    mobile: Optional[str] = None
    password: str

class VisitorIn(BaseModel):
    full_name: str
    mobile: str
    business_name: Optional[str] = ""
    industry: Optional[str] = ""
    city: Optional[str] = ""
    referred_by: Optional[str] = ""
    email: Optional[str] = ""
    is_lvb_member: Optional[bool] = False
    lvb_chapter: Optional[str] = ""
    photo_url: Optional[str] = ""

class VisitorOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    full_name: str
    mobile: str
    business_name: str = ""
    industry: str = ""
    city: str = ""
    referred_by: str = ""
    email: str = ""
    is_lvb_member: bool = False
    lvb_chapter: str = ""
    photo_url: str = ""
    qr_id: str
    attended: bool = False
    attended_at: Optional[str] = None
    created_at: str

class MemberNumberIn(BaseModel):
    mobile: str
    note: Optional[str] = ""

class CatalogueItem(BaseModel):
    name: Optional[str] = ""
    description: Optional[str] = ""
    image_url: Optional[str] = ""

class Testimonial(BaseModel):
    name: Optional[str] = ""
    role: Optional[str] = ""
    text: Optional[str] = ""

class CustomLink(BaseModel):
    label: Optional[str] = ""
    url: Optional[str] = ""

class ExhibitorRegisterIn(BaseModel):
    mobile: str
    password: str
    member_name: str
    business_name: str
    category: str
    position: Optional[str] = ""
    whatsapp: Optional[str] = ""
    email: Optional[str] = ""
    description: Optional[str] = ""
    products_services: Optional[str] = ""
    instagram: Optional[str] = ""
    facebook: Optional[str] = ""
    linkedin: Optional[str] = ""
    website: Optional[str] = ""
    address: Optional[str] = ""
    maps_link: Optional[str] = ""
    shop_address: Optional[str] = ""
    shop_maps_link: Optional[str] = ""
    logo_url: Optional[str] = ""
    banner_url: Optional[str] = ""
    profile_photo_url: Optional[str] = ""
    photo_focus_x: Optional[float] = Field(default=0.5, ge=0.0, le=1.0)
    photo_focus_y: Optional[float] = Field(default=0.35, ge=0.0, le=1.0)
    photo_zoom: Optional[float] = Field(default=1.0, ge=1.0, le=3.0)
    catalogue_pdf_url: Optional[str] = ""
    catalogue_gallery: Optional[List[CatalogueItem]] = None
    testimonials: Optional[List[Testimonial]] = None
    custom_links: Optional[List[CustomLink]] = None

class ExhibitorUpdateIn(BaseModel):
    member_name: Optional[str] = None
    business_name: Optional[str] = None
    category: Optional[str] = None
    position: Optional[str] = None
    whatsapp: Optional[str] = None
    email: Optional[str] = None
    description: Optional[str] = None
    products_services: Optional[str] = None
    instagram: Optional[str] = None
    facebook: Optional[str] = None
    linkedin: Optional[str] = None
    website: Optional[str] = None
    address: Optional[str] = None
    maps_link: Optional[str] = None
    shop_address: Optional[str] = None
    shop_maps_link: Optional[str] = None
    logo_url: Optional[str] = None
    banner_url: Optional[str] = None
    profile_photo_url: Optional[str] = None
    photo_focus_x: Optional[float] = Field(default=None, ge=0.0, le=1.0)
    photo_focus_y: Optional[float] = Field(default=None, ge=0.0, le=1.0)
    photo_zoom: Optional[float] = Field(default=None, ge=1.0, le=3.0)
    catalogue_pdf_url: Optional[str] = None
    catalogue_gallery: Optional[List[CatalogueItem]] = None
    testimonials: Optional[List[Testimonial]] = None
    custom_links: Optional[List[CustomLink]] = None

class PasswordResetIn(BaseModel):
    new_password: str

class SponsorAdIn(BaseModel):
    name: str
    placement: Literal["popup", "inline", "footer", "featured"] = "inline"
    media_type: Literal["image", "video"] = "image"
    media_url: str
    link: Optional[str] = ""
    active: bool = True
    order: int = 0

class EventSettingsIn(BaseModel):
    event_name: Optional[str] = None
    venue: Optional[str] = None
    venue_address: Optional[str] = None
    maps_link: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    visitor_registration_open: Optional[bool] = None
    exhibitor_registration_open: Optional[bool] = None
    exhibitor_limit: Optional[int] = None
    whatsapp_template_visitor: Optional[str] = None
    whatsapp_template_exhibitor: Optional[str] = None
    bizchat_vendor_uid: Optional[str] = None
    bizchat_token: Optional[str] = None
    bizchat_template_visitor: Optional[str] = None
    bizchat_template_exhibitor: Optional[str] = None
    bizchat_template_language: Optional[str] = None
    bizchat_from_phone_id: Optional[str] = None
    gate_code: Optional[str] = None

# ---------- Sanitize helpers ----------
EXHIBITOR_PUBLIC_FIELDS = {
    "id", "slug", "mobile", "member_name", "business_name", "category", "position", "whatsapp", "email",
    "description", "products_services", "instagram", "facebook", "linkedin", "website",
    "address", "maps_link", "shop_address", "shop_maps_link",
    "logo_url", "banner_url", "profile_photo_url",
    "photo_focus_x", "photo_focus_y", "photo_zoom",
    "catalogue_pdf_url", "catalogue_gallery", "testimonials", "custom_links",
    "approved", "featured", "hidden", "created_at"
}

# Short, URL-safe IDs for public digital cards (6 chars, ~56 bits of entropy → safe for our scale)
_SLUG_ALPHABET = "abcdefghijkmnopqrstuvwxyz23456789"  # no 0/1/l/o to keep it readable on print
def _make_slug(n: int = 6) -> str:
    return "".join(secrets.choice(_SLUG_ALPHABET) for _ in range(n))

async def _generate_unique_slug() -> str:
    for _ in range(20):
        s = _make_slug()
        if not await db.exhibitors.find_one({"slug": s}, {"_id": 1}):
            return s
    # Fallback to longer slug if astronomically unlucky
    return _make_slug(8)

_EXHIBITOR_LIST_DEFAULTS = {"catalogue_gallery", "testimonials", "custom_links"}
_EXHIBITOR_STR_DEFAULTS = {"catalogue_pdf_url", "shop_address", "shop_maps_link"}

def public_exhibitor(doc: dict) -> dict:
    out = {k: doc.get(k) for k in EXHIBITOR_PUBLIC_FIELDS if k in doc}
    # Stable schema: always expose digital-card fields with safe defaults
    for k in _EXHIBITOR_LIST_DEFAULTS:
        out[k] = doc.get(k) or []
    for k in _EXHIBITOR_STR_DEFAULTS:
        out[k] = doc.get(k) or ""
    return out

def admin_exhibitor(doc: dict) -> dict:
    d = dict(doc)
    d.pop("_id", None)
    d.pop("password_hash", None)
    return d

# ---------- Auth Endpoints ----------
@api.post("/auth/login")
async def login(data: LoginIn):
    # Admin login via email
    if data.email:
        admin = await db.admins.find_one({"email": data.email.lower()}, {"_id": 0})
        if not admin or not verify_password(data.password, admin["password_hash"]):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        token = create_token(admin["id"], "admin", {"email": admin["email"]})
        return {"token": token, "role": "admin", "user": {"id": admin["id"], "email": admin["email"], "name": admin.get("name", "Admin")}}
    # Exhibitor login via mobile
    if data.mobile:
        mobile = normalize_mobile(data.mobile)
        ex = await db.exhibitors.find_one({"mobile": mobile}, {"_id": 0})
        if not ex or not ex.get("password_hash") or not verify_password(data.password, ex["password_hash"]):
            raise HTTPException(status_code=401, detail="Invalid mobile or password")
        token = create_token(ex["id"], "exhibitor", {"mobile": ex["mobile"]})
        return {"token": token, "role": "exhibitor", "user": public_exhibitor(ex)}
    raise HTTPException(status_code=400, detail="Provide email or mobile")

@api.get("/auth/me")
async def me(user: dict = Depends(get_current)):
    if user["role"] == "admin":
        a = await db.admins.find_one({"id": user["sub"]}, {"_id": 0, "password_hash": 0})
        return {"role": "admin", "user": a}
    if user["role"] == "gate":
        return {"role": "gate", "user": {"id": "gate", "name": "Gate Staff"}}
    ex = await db.exhibitors.find_one({"id": user["sub"]}, {"_id": 0, "password_hash": 0})
    if not ex:
        raise HTTPException(status_code=401, detail="Not found")
    return {"role": "exhibitor", "user": public_exhibitor(ex)}

# ---------- Gate Staff Login (QR scanner only) ----------
class GateLoginIn(BaseModel):
    code: str

@api.post("/gate/login")
async def gate_login(data: GateLoginIn):
    s = await get_settings()
    expected = (s.get("gate_code") or "").strip()
    given = (data.code or "").strip()
    if not expected:
        raise HTTPException(status_code=400, detail="Gate code not configured. Ask admin to set it in Settings.")
    if not secrets.compare_digest(expected, given):
        raise HTTPException(status_code=401, detail="Invalid gate code")
    # Long-lived token (30 days) so a single login covers full event days
    payload = {
        "sub": "gate",
        "role": "gate",
        "exp": datetime.now(timezone.utc) + timedelta(days=30),
        "iat": datetime.now(timezone.utc),
    }
    token = jwt.encode(payload, jwt_secret(), algorithm=JWT_ALGORITHM)
    return {"token": token, "user": {"role": "gate", "name": "Gate Staff"}}

# ---------- File Upload ----------
@api.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in {"png", "jpg", "jpeg", "webp", "gif", "mp4", "webm", "pdf"}:
        raise HTTPException(status_code=400, detail="Unsupported file type")
    file_id = f"{uuid.uuid4().hex}.{ext}"
    path = UPLOAD_DIR / file_id
    content = await file.read()
    max_mb = 20 if ext == "pdf" else 8
    if len(content) > max_mb * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"File too large (max {max_mb}MB)")
    with open(path, "wb") as f:
        f.write(content)
    return {"url": f"/api/uploads/{file_id}", "filename": file_id}

# ---------- Public Digital Card ----------
@api.get("/c/{slug}")
async def public_card(slug: str):
    ex = await db.exhibitors.find_one({"slug": slug, "hidden": {"$ne": True}}, {"_id": 0, "password_hash": 0})
    if not ex:
        raise HTTPException(status_code=404, detail="Card not found")
    return public_exhibitor(ex)

@api.get("/c/{slug}/vcard")
async def public_card_vcard(slug: str):
    ex = await db.exhibitors.find_one({"slug": slug, "hidden": {"$ne": True}}, {"_id": 0, "password_hash": 0})
    if not ex:
        raise HTTPException(status_code=404, detail="Card not found")
    fn = (ex.get("member_name") or "").strip() or (ex.get("business_name") or "Contact")
    org = (ex.get("business_name") or "").strip()
    title = (ex.get("position") or "").strip()
    tel = (ex.get("whatsapp") or ex.get("mobile") or "").strip()
    email = (ex.get("email") or "").strip()
    url = (ex.get("website") or "").strip()
    address = (ex.get("shop_address") or ex.get("address") or "").strip()
    note = (ex.get("description") or "").strip()
    socials = []
    for k in ("instagram", "facebook", "linkedin"):
        v = (ex.get(k) or "").strip()
        if v:
            socials.append(("X-SOCIALPROFILE;TYPE=" + k, v))
    parts = ["BEGIN:VCARD", "VERSION:3.0", f"FN:{fn}"]
    if org:
        parts.append(f"ORG:{org}")
    if title:
        parts.append(f"TITLE:{title}")
    if tel:
        parts.append(f"TEL;TYPE=CELL,VOICE:+91{tel}" if tel.isdigit() and len(tel) == 10 else f"TEL;TYPE=CELL,VOICE:{tel}")
    if email:
        parts.append(f"EMAIL;TYPE=INTERNET:{email}")
    if url:
        parts.append(f"URL:{url}")
    if address:
        parts.append(f"ADR;TYPE=WORK:;;{address};;;;")
    if note:
        parts.append(f"NOTE:{note}")
    for k, v in socials:
        parts.append(f"{k}:{v}")
    parts.append("END:VCARD")
    body = "\r\n".join(parts) + "\r\n"
    fname = (org or fn or "contact").replace(" ", "_").lower()
    return StreamingResponse(
        io.BytesIO(body.encode("utf-8")),
        media_type="text/vcard",
        headers={"Content-Disposition": f'attachment; filename="{fname}.vcf"'},
    )

@api.get("/c/{slug}/qr.png")
async def public_card_qr(slug: str, request: Request):
    ex = await db.exhibitors.find_one({"slug": slug, "hidden": {"$ne": True}}, {"_id": 0, "slug": 1, "business_name": 1})
    if not ex:
        raise HTTPException(status_code=404, detail="Card not found")
    base = os.environ.get("PUBLIC_BASE_URL", "").rstrip("/") or f"{request.url.scheme}://{request.url.netloc}"
    url = f"{base}/c/{slug}"
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=20, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#1f1f27", back_color="#ffffff").convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    buf.seek(0)
    fname = (ex.get("business_name") or slug).replace(" ", "_").lower()
    return StreamingResponse(buf, media_type="image/png", headers={
        "Content-Disposition": f'inline; filename="{fname}-card-qr.png"',
        "Cache-Control": "public, max-age=3600",
    })

# ---------- Public Settings ----------
async def get_settings() -> dict:
    s = await db.settings.find_one({"id": "main"}, {"_id": 0})
    if not s:
        s = {
            "id": "main",
            "event_name": "Rama Bazaar 1.0",
            "venue": "TBA",
            "venue_address": "",
            "maps_link": "",
            "start_date": "",
            "end_date": "",
            "visitor_registration_open": True,
            "exhibitor_registration_open": True,
            "exhibitor_limit": int(os.environ.get("EXHIBITOR_REGISTRATION_LIMIT", "100")),
            "whatsapp_template_visitor": "Hello {name}, your Rama Bazaar 1.0 registration is confirmed. Your QR is attached. — Team Rama",
            "whatsapp_template_exhibitor": "Welcome {name}, your Rama Bazaar 1.0 exhibitor account is created. Login at the portal.",
            "bizchat_vendor_uid": os.environ.get("BIZCHAT_VENDOR_UID", ""),
            "bizchat_token": os.environ.get("BIZCHAT_TOKEN", ""),
            "bizchat_template_visitor": "",
            "bizchat_template_exhibitor": "",
            "bizchat_template_language": "en",
            "bizchat_from_phone_id": "",
            "updated_at": now_iso(),
        }
        await db.settings.insert_one(dict(s))
    return s

@api.get("/settings")
async def public_settings():
    s = await get_settings()
    # Don't leak secrets publicly
    safe = {k: v for k, v in s.items() if k not in {"bizchat_token", "bizchat_vendor_uid", "bizchat_from_phone_id"}}
    return safe

@api.get("/admin/settings")
async def admin_get_settings(_: dict = Depends(require_admin)):
    return await get_settings()

@api.put("/admin/settings")
async def admin_update_settings(data: EventSettingsIn, _: dict = Depends(require_admin)):
    upd = {k: v for k, v in data.model_dump(exclude_none=True).items()}
    upd["updated_at"] = now_iso()
    await db.settings.update_one({"id": "main"}, {"$set": upd}, upsert=True)
    return await get_settings()

# ---------- Visitors ----------
@api.post("/visitors/register")
async def visitor_register(data: VisitorIn):
    s = await get_settings()
    if not s.get("visitor_registration_open", True):
        raise HTTPException(status_code=403, detail="Visitor registration is closed")
    mobile = normalize_mobile(data.mobile)
    if len(mobile) != 10:
        raise HTTPException(status_code=400, detail="Invalid mobile number")
    existing = await db.visitors.find_one({"mobile": mobile}, {"_id": 0})
    if existing:
        return {"already_registered": True, "visitor": existing}
    qr_id = uuid.uuid4().hex
    doc = {
        "id": uuid.uuid4().hex,
        "full_name": data.full_name.strip(),
        "mobile": mobile,
        "business_name": (data.business_name or "").strip(),
        "industry": (data.industry or "").strip(),
        "city": (data.city or "").strip(),
        "referred_by": (data.referred_by or "").strip(),
        "email": (data.email or "").strip(),
        "is_lvb_member": bool(data.is_lvb_member),
        "lvb_chapter": (data.lvb_chapter or "").strip() if data.is_lvb_member else "",
        "photo_url": (data.photo_url or "").strip(),
        "qr_id": qr_id,
        "attended": False,
        "attended_at": None,
        "created_at": now_iso(),
    }
    await db.visitors.insert_one(dict(doc))
    return {"already_registered": False, "visitor": doc}

@api.post("/visitors/retrieve")
async def visitor_retrieve(payload: dict):
    mobile = normalize_mobile(payload.get("mobile", ""))
    v = await db.visitors.find_one({"mobile": mobile}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="No registration found for this mobile number")
    return {"visitor": v}

@api.get("/visitors/by-qr/{qr_id}")
async def visitor_by_qr(qr_id: str):
    v = await db.visitors.find_one({"qr_id": qr_id}, {"_id": 0, "password_hash": 0})
    if not v:
        raise HTTPException(status_code=404, detail="QR not found")
    return {"visitor": v}


@api.head("/visitors/qr/{qr_id}.png")
async def visitor_qr_image_head(qr_id: str):
    """HEAD support for media downloaders (BizChat/WhatsApp validate URL before fetch)."""
    v = await db.visitors.find_one({"qr_id": qr_id}, {"_id": 0, "qr_id": 1})
    if not v:
        raise HTTPException(status_code=404, detail="QR not found")
    return JSONResponse(
        content=None,
        headers={
            "Content-Type": "image/png",
            "Cache-Control": "public, max-age=600",
            "Access-Control-Allow-Origin": "*",
        },
    )


@api.get("/visitors/qr/{qr_id}.png")
async def visitor_qr_image(qr_id: str, plain: int = 0):
    v = await db.visitors.find_one({"qr_id": qr_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="QR not found")
    payload = json.dumps({"qr": qr_id, "name": v["full_name"], "mobile": v["mobile"]})
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=40, border=2)
    qr.add_data(payload)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="#1f1f27", back_color="#ffffff").convert("RGB")

    # Plain mode — return just the QR (used by the on-page preview card).
    if plain:
        from PIL import Image as _Image
        qr_img = qr_img.resize((800, 800), _Image.LANCZOS)
        buf = io.BytesIO()
        qr_img.save(buf, format="PNG", optimize=False, compress_level=3)
        buf.seek(0)
        return StreamingResponse(buf, media_type="image/png", headers={
            "Cache-Control": "public, max-age=86400",
        })

    # Compose branded poster: 1080×1900 portrait — mirrors the on-page ticket card
    from PIL import Image, ImageDraw, ImageFilter
    W, H = 1080, 1900
    bg = Image.new("RGB", (W, H), "#f5efe1")
    draw = ImageDraw.Draw(bg)

    brand = ROOT_DIR / "assets" / "brand"

    def _paste(path: Path, max_w: int, max_h: int, cx: int, cy: int):
        if not path.exists():
            return None
        try:
            im = Image.open(path).convert("RGBA")
            ratio = min(max_w / im.width, max_h / im.height)
            new_w, new_h = int(im.width * ratio), int(im.height * ratio)
            im = im.resize((new_w, new_h), Image.LANCZOS)
            x, y = cx - new_w // 2, cy - new_h // 2
            bg.paste(im, (x, y), im if im.mode == "RGBA" else None)
            return (x, y, new_w, new_h)
        except Exception as e:
            logger.warning(f"poster logo paste failed for {path.name}: {e}")
            return None

    # Fonts
    eyebrow_md    = _cinzel(26)
    visitor_lbl_f = _cinzel(30)
    name_italic_f = _truetype(64, italic=True)
    mobile_f      = _truetype(34, bold=False)
    chip_f        = _cinzel(22)
    short_id_f    = _cinzel(22)
    info_f        = _truetype(30, bold=False)
    micro_f       = _cinzel(18)

    # ============================================================
    # 1) POWERED BY header — gold-flanked eyebrow + rounded navy band
    # ============================================================
    # Eyebrow with side dashes
    draw.text((W // 2, 100), "POWERED BY", font=eyebrow_md, fill="#b2873d", anchor="mt")
    draw.line([(W // 2 - 290, 118), (W // 2 - 130, 118)], fill="#b2873d", width=1)
    draw.line([(W // 2 + 130, 118), (W // 2 + 290, 118)], fill="#b2873d", width=1)

    # Navy band with rounded corners + soft shadow + gold inner hairline
    band_x0, band_y0, band_x1, band_y1 = 80, 170, W - 80, 380
    # Shadow
    shadow = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    sd = ImageDraw.Draw(shadow)
    sd.rounded_rectangle([(band_x0 + 6, band_y0 + 10), (band_x1 + 6, band_y1 + 10)],
                         radius=18, fill=(27, 25, 75, 70))
    shadow = shadow.filter(ImageFilter.GaussianBlur(6))
    bg.paste(shadow, (0, 0), shadow)
    draw.rounded_rectangle([(band_x0, band_y0), (band_x1, band_y1)], radius=18, fill="#1B194B")
    draw.rounded_rectangle([(band_x0 + 6, band_y0 + 6), (band_x1 - 6, band_y1 - 6)],
                           radius=14, outline="#c19b30", width=1)
    _paste(brand / "coco-salons.jpg",
           max_w=band_x1 - band_x0 - 120, max_h=130,
           cx=W // 2, cy=(band_y0 + band_y1) // 2)

    # ============================================================
    # 2) MAIN TICKET CARD — white, rounded, gold hairline + shadow
    # ============================================================
    card_x0, card_x1 = 80, W - 80
    card_y0, card_y1 = 420, 1790
    shadow = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    sd = ImageDraw.Draw(shadow)
    sd.rounded_rectangle([(card_x0 + 4, card_y0 + 14), (card_x1 + 4, card_y1 + 14)],
                         radius=28, fill=(27, 25, 75, 60))
    shadow = shadow.filter(ImageFilter.GaussianBlur(10))
    bg.paste(shadow, (0, 0), shadow)
    draw.rounded_rectangle([(card_x0, card_y0), (card_x1, card_y1)],
                           radius=28, fill="#ffffff", outline="#d8bc84", width=1)
    # Slim gold band at the top edge of the card
    draw.rounded_rectangle([(card_x0, card_y0), (card_x1, card_y0 + 8)],
                           radius=8, fill="#b2873d")

    # VISITOR PASS eyebrow
    draw.text((W // 2, card_y0 + 50), "VISITOR PASS",
              font=visitor_lbl_f, fill="#b2873d", anchor="mt")

    # Official Rama Bazaar lockup — large, hero placement
    _paste(brand / "rama-bazaar-lockup.png",
           max_w=620, max_h=420, cx=W // 2, cy=card_y0 + 310)

    # Visitor name (italic Playfair) + mobile — pulled up close to the lockup
    name = (v.get("full_name") or "").strip()
    name_y = card_y0 + 540
    draw.text((W // 2, name_y), name[:30],
              font=name_italic_f, fill="#1B194B", anchor="mt")
    draw.text((W // 2, name_y + 84), f"+91 {v.get('mobile', '')}",
              font=mobile_f, fill="#3b3b46", anchor="mt")

    # LVB chapter pill (rounded gold border, cream fill, gold text)
    pill_y_center = name_y + 168
    if v.get("is_lvb_member") and v.get("lvb_chapter"):
        chip_text = f"LVB  ·  {v['lvb_chapter'].upper()}  CHAPTER"
        bbox = draw.textbbox((0, 0), chip_text, font=chip_f, anchor="lt")
        cw = bbox[2] - bbox[0] + 80
        ch = 56
        cx0 = (W - cw) // 2
        cy0 = pill_y_center - ch // 2
        draw.rounded_rectangle([(cx0, cy0), (cx0 + cw, cy0 + ch)],
                               radius=ch // 2, fill="#fbf6e8", outline="#b2873d", width=1)
        draw.text((W // 2, cy0 + ch // 2), chip_text,
                  font=chip_f, fill="#b2873d", anchor="mm")

    # --- Perforation divider ---
    perf_y = name_y + 215
    notch_r = 18
    draw.ellipse([(card_x0 - notch_r, perf_y - notch_r),
                  (card_x0 + notch_r, perf_y + notch_r)],
                 fill="#f5efe1", outline="#d8bc84", width=1)
    draw.ellipse([(card_x1 - notch_r, perf_y - notch_r),
                  (card_x1 + notch_r, perf_y + notch_r)],
                 fill="#f5efe1", outline="#d8bc84", width=1)
    dash_x0 = card_x0 + notch_r + 10
    dash_x1 = card_x1 - notch_r - 10
    x = dash_x0
    while x < dash_x1:
        x_end = min(x + 14, dash_x1)
        draw.line([(x, perf_y), (x_end, perf_y)], fill="#b2873d", width=2)
        x += 26

    # --- QR sub-card (cream bg, gold border) — contains QR + scan + date/venue ---
    qr_card_x0, qr_card_x1 = card_x0 + 100, card_x1 - 100
    qr_card_y0 = perf_y + 30
    qr_card_y1 = qr_card_y0 + 560
    draw.rounded_rectangle([(qr_card_x0, qr_card_y0), (qr_card_x1, qr_card_y1)],
                           radius=20, fill="#fbf8f0", outline="#d8bc84", width=1)
    qr_size = 320
    qr_img_r = qr_img.resize((qr_size, qr_size), Image.LANCZOS)
    bg.paste(qr_img_r, ((W - qr_size) // 2, qr_card_y0 + 28))

    # Scan caption inside QR sub-card, below the QR
    scan_y = qr_card_y0 + 375
    draw.text((W // 2, scan_y), f"SCAN AT VENUE  ·  {qr_id[:6].upper()}",
              font=short_id_f, fill="#b2873d", anchor="mt")

    # Date row + Venue row — each on its own centered line so long venue names fit
    try:
        s = await get_settings()
        sd_, ed_ = s.get("start_date", ""), s.get("end_date", "")
        venue = s.get("venue", "")
        date_str_disp = sd_ if (not ed_ or sd_ == ed_) else f"{sd_} – {ed_}"

        def _draw_centered_row(row_y: int, icon: str, text: str):
            if not text:
                return
            tb = draw.textbbox((0, 0), text, font=info_f, anchor="lt")
            tw = tb[2] - tb[0]
            icon_w = 26  # icon + gap
            total = icon_w + tw
            start_x = (W - total) // 2
            icon_cx = start_x + 9
            if icon == "cal":
                draw.rectangle([(icon_cx - 10, row_y - 12), (icon_cx + 10, row_y + 12)],
                               outline="#b2873d", width=2)
                draw.line([(icon_cx - 10, row_y - 4), (icon_cx + 10, row_y - 4)],
                          fill="#b2873d", width=2)
            elif icon == "pin":
                draw.ellipse([(icon_cx - 9, row_y - 12), (icon_cx + 9, row_y + 6)],
                             outline="#b2873d", width=2)
                draw.polygon([(icon_cx - 5, row_y + 4), (icon_cx + 5, row_y + 4),
                              (icon_cx, row_y + 14)], fill="#b2873d")
            draw.text((start_x + icon_w, row_y), text,
                      font=info_f, fill="#3b3b46", anchor="lm")

        date_row_y = qr_card_y0 + 445
        venue_row_y = qr_card_y0 + 500
        _draw_centered_row(date_row_y, "cal", date_str_disp)
        _draw_centered_row(venue_row_y, "pin", venue)
    except Exception:
        pass

    # ============================================================
    # 3) TECHNOLOGY PARTNER — OUTSIDE the card, in cream zone
    # ============================================================
    tp_y = card_y1 + 30
    draw.line([(W // 2 - 180, tp_y), (W // 2 - 90, tp_y)], fill="#d8bc84", width=1)
    draw.line([(W // 2 + 90, tp_y), (W // 2 + 180, tp_y)], fill="#d8bc84", width=1)
    draw.text((W // 2, tp_y + 10), "TECHNOLOGY PARTNER",
              font=micro_f, fill="#7a7868", anchor="mt")
    _paste(brand / "rxt.png", max_w=200, max_h=28, cx=W // 2, cy=tp_y + 56)

    buf = io.BytesIO()
    bg.save(buf, format="PNG", optimize=False, compress_level=3)
    buf.seek(0)
    headers = {"Content-Disposition": f'inline; filename="rama-bazaar-{qr_id}.png"', "Cache-Control": "public, max-age=86400"}
    return StreamingResponse(buf, media_type="image/png", headers=headers)

# ---------- Attendance ----------
@api.post("/attendance/scan")
async def attendance_scan(payload: dict, _: dict = Depends(require_gate_or_admin)):
    qr_id = payload.get("qr_id", "").strip()
    # Allow JSON QR payloads too
    if qr_id.startswith("{"):
        try:
            qr_id = json.loads(qr_id).get("qr", "")
        except Exception:
            pass
    v = await db.visitors.find_one({"qr_id": qr_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Invalid QR code")
    if v.get("attended"):
        return {"already": True, "visitor": v}
    await db.visitors.update_one({"qr_id": qr_id}, {"$set": {"attended": True, "attended_at": now_iso()}})
    v["attended"] = True
    v["attended_at"] = now_iso()
    return {"already": False, "visitor": v}

@api.post("/attendance/manual")
async def attendance_manual(payload: dict, _: dict = Depends(require_gate_or_admin)):
    mobile = normalize_mobile(payload.get("mobile", ""))
    v = await db.visitors.find_one({"mobile": mobile}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Visitor not found")
    if v.get("attended"):
        return {"already": True, "visitor": v}
    await db.visitors.update_one({"mobile": mobile}, {"$set": {"attended": True, "attended_at": now_iso()}})
    v["attended"] = True
    v["attended_at"] = now_iso()
    return {"already": False, "visitor": v}

@api.get("/attendance/stats")
async def attendance_stats(_: dict = Depends(require_gate_or_admin)):
    total = await db.visitors.count_documents({})
    present = await db.visitors.count_documents({"attended": True})
    return {"total": total, "present": present, "pending": total - present}

# ---------- Allowed Member Numbers ----------
@api.post("/admin/members")
async def add_member(data: MemberNumberIn, _: dict = Depends(require_admin)):
    mobile = normalize_mobile(data.mobile)
    if len(mobile) != 10:
        raise HTTPException(status_code=400, detail="Invalid mobile number")
    existing = await db.allowed_members.find_one({"mobile": mobile})
    if existing:
        return {"already": True}
    await db.allowed_members.insert_one({"id": uuid.uuid4().hex, "mobile": mobile, "note": data.note or "", "created_at": now_iso()})
    return {"ok": True}

@api.post("/admin/members/bulk")
async def add_members_bulk(file: UploadFile = File(...), _: dict = Depends(require_admin)):
    content = (await file.read()).decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(content))
    count = 0
    skipped = 0
    for row in reader:
        if not row:
            continue
        mobile = normalize_mobile(row[0])
        if len(mobile) != 10:
            skipped += 1
            continue
        existing = await db.allowed_members.find_one({"mobile": mobile})
        if existing:
            skipped += 1
            continue
        note = row[1] if len(row) > 1 else ""
        await db.allowed_members.insert_one({"id": uuid.uuid4().hex, "mobile": mobile, "note": note, "created_at": now_iso()})
        count += 1
    return {"added": count, "skipped": skipped}

@api.get("/admin/members")
async def list_members(_: dict = Depends(require_admin)):
    items = await db.allowed_members.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return items

@api.delete("/admin/members/{member_id}")
async def delete_member(member_id: str, _: dict = Depends(require_admin)):
    await db.allowed_members.delete_one({"id": member_id})
    return {"ok": True}

@api.post("/exhibitors/check-eligibility")
async def check_eligibility(payload: dict):
    mobile = normalize_mobile(payload.get("mobile", ""))
    if len(mobile) != 10:
        raise HTTPException(status_code=400, detail="Invalid mobile number")
    s = await get_settings()
    count = await db.exhibitors.count_documents({})
    allowed = await db.allowed_members.find_one({"mobile": mobile})
    already = await db.exhibitors.find_one({"mobile": mobile})
    if already:
        raise HTTPException(status_code=409, detail="This mobile is already registered as an exhibitor")
    if not allowed:
        raise HTTPException(status_code=403, detail="This number is not eligible for Rama Bazaar exhibitor registration.")
    if not s.get("exhibitor_registration_open", True):
        raise HTTPException(status_code=403, detail="Exhibitor registration is closed")
    limit = int(s.get("exhibitor_limit", 100))
    if count >= limit:
        raise HTTPException(status_code=403, detail="Exhibitor registration limit reached")
    return {"eligible": True, "slots_remaining": max(0, limit - count)}

@api.post("/exhibitors/register")
async def exhibitor_register(data: ExhibitorRegisterIn):
    mobile = normalize_mobile(data.mobile)
    s = await get_settings()
    count = await db.exhibitors.count_documents({})
    allowed = await db.allowed_members.find_one({"mobile": mobile})
    if not allowed:
        raise HTTPException(status_code=403, detail="This number is not eligible for Rama Bazaar exhibitor registration.")
    if not s.get("exhibitor_registration_open", True):
        raise HTTPException(status_code=403, detail="Exhibitor registration is closed")
    limit = int(s.get("exhibitor_limit", 100))
    if count >= limit:
        raise HTTPException(status_code=403, detail="Exhibitor registration limit reached")
    if await db.exhibitors.find_one({"mobile": mobile}):
        raise HTTPException(status_code=409, detail="This mobile is already registered")
    doc = data.model_dump()
    doc["mobile"] = mobile
    doc["whatsapp"] = normalize_mobile(doc.get("whatsapp", "")) or mobile
    doc["id"] = uuid.uuid4().hex
    doc["slug"] = await _generate_unique_slug()
    doc["password_hash"] = hash_password(data.password)
    doc.pop("password", None)
    doc["approved"] = False
    doc["featured"] = False
    doc["hidden"] = False
    doc["created_at"] = now_iso()
    await db.exhibitors.insert_one(dict(doc))
    token = create_token(doc["id"], "exhibitor", {"mobile": mobile})
    return {"token": token, "role": "exhibitor", "user": public_exhibitor(doc)}

@api.put("/exhibitors/me")
async def update_self(data: ExhibitorUpdateIn, user: dict = Depends(require_exhibitor)):
    upd = {k: v for k, v in data.model_dump(exclude_none=True).items()}
    if not upd:
        raise HTTPException(status_code=400, detail="Nothing to update")
    if "whatsapp" in upd:
        upd["whatsapp"] = normalize_mobile(upd["whatsapp"])
    upd["updated_at"] = now_iso()
    await db.exhibitors.update_one({"id": user["sub"]}, {"$set": upd})
    ex = await db.exhibitors.find_one({"id": user["sub"]}, {"_id": 0})
    return public_exhibitor(ex)

@api.post("/exhibitors/me/password")
async def change_password(payload: dict, user: dict = Depends(require_exhibitor)):
    new_pw = payload.get("new_password", "")
    if len(new_pw) < 6:
        raise HTTPException(status_code=400, detail="Password too short")
    await db.exhibitors.update_one({"id": user["sub"]}, {"$set": {"password_hash": hash_password(new_pw)}})
    return {"ok": True}

# ---------- Exhibitor Social Post (Generated Share Image) ----------
SOCIAL_TEMPLATE_PATH = ROOT_DIR / "assets" / "social-template.png"
# Render at half resolution (1500x1500) for ~10x speedup. Still crisp on WhatsApp/Instagram.
RENDER_SCALE = 0.5
RENDER_W = int(3000 * RENDER_SCALE)
RENDER_H = int(3000 * RENDER_SCALE)
# Coordinates scaled down from the original 3000×3000 layout:
SILHOUETTE_BBOX = (int(0 * RENDER_SCALE), int(25 * RENDER_SCALE), int(965 * RENDER_SCALE), int(2695 * RENDER_SCALE))
TEXT_BOX = (int(1010 * RENDER_SCALE), int(2070 * RENDER_SCALE), int(2580 * RENDER_SCALE), int(2580 * RENDER_SCALE))
NAVY_INK = (27, 25, 75)
GOLD_INK = (178, 135, 61)

# Cache the resized template once at startup (avoids re-decoding 8MB PNG per request)
_TEMPLATE_CACHE = {"img": None, "mtime": None}
def _get_template():
    from PIL import Image
    if not SOCIAL_TEMPLATE_PATH.exists():
        return None
    mtime = SOCIAL_TEMPLATE_PATH.stat().st_mtime
    if _TEMPLATE_CACHE["img"] is None or _TEMPLATE_CACHE["mtime"] != mtime:
        t = Image.open(SOCIAL_TEMPLATE_PATH).convert("RGBA")
        if t.size != (RENDER_W, RENDER_H):
            t = t.resize((RENDER_W, RENDER_H), Image.LANCZOS)
        _TEMPLATE_CACHE["img"] = t
        _TEMPLATE_CACHE["mtime"] = mtime
    return _TEMPLATE_CACHE["img"]

# Generated-post cache (in-memory, up to 200 entries) keyed by exhibitor + framing + photo mtime
_POST_CACHE: dict = {}
_POST_CACHE_ORDER: list = []
_POST_CACHE_MAX = 200

def _post_cache_get(key: str):
    return _POST_CACHE.get(key)

def _post_cache_set(key: str, png: bytes):
    if key in _POST_CACHE:
        _POST_CACHE_ORDER.remove(key)
    elif len(_POST_CACHE_ORDER) >= _POST_CACHE_MAX:
        evict = _POST_CACHE_ORDER.pop(0)
        _POST_CACHE.pop(evict, None)
    _POST_CACHE[key] = png
    _POST_CACHE_ORDER.append(key)

def _truetype(size: int, bold: bool = False, italic: bool = False):
    """Resolve a luxury serif font with bold/italic variants.
    Playfair Display ships with the app for the premium aesthetic."""
    fonts_dir = ROOT_DIR / "assets" / "fonts"
    bold_italic = bold and italic
    candidates: list[str] = []
    # Bundled Playfair Display (preferred for luxury)
    if bold_italic:
        candidates += [str(fonts_dir / "PlayfairDisplay-Italic.ttf")]
    elif bold:
        candidates += [str(fonts_dir / "PlayfairDisplay-Bold.ttf")]
    elif italic:
        candidates += [str(fonts_dir / "PlayfairDisplay-Italic.ttf")]
    else:
        candidates += [str(fonts_dir / "PlayfairDisplay-Bold.ttf")]
    # System fallbacks
    if bold_italic:
        candidates += [
            "/usr/share/fonts/truetype/liberation/LiberationSerif-BoldItalic.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSerifBoldItalic.ttf",
        ]
    elif bold:
        candidates += [
            "/usr/share/fonts/truetype/liberation/LiberationSerif-Bold.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSerifBold.ttf",
        ]
    elif italic:
        candidates += [
            "/usr/share/fonts/truetype/liberation/LiberationSerif-Italic.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSerifItalic.ttf",
        ]
    else:
        candidates += [
            "/usr/share/fonts/truetype/liberation/LiberationSerif-Regular.ttf",
            "/usr/share/fonts/truetype/freefont/FreeSerif.ttf",
        ]
    for p in candidates:
        if os.path.exists(p):
            try:
                from PIL import ImageFont
                return ImageFont.truetype(p, size)
            except Exception:
                pass
    from PIL import ImageFont
    return ImageFont.load_default()


def _cinzel(size: int):
    """Cinzel — uppercase Roman serif for the category line."""
    p = str(ROOT_DIR / "assets" / "fonts" / "Cinzel-Bold.ttf")
    if os.path.exists(p):
        try:
            from PIL import ImageFont
            return ImageFont.truetype(p, size)
        except Exception:
            pass
    return _truetype(size, bold=True)

def _fit_text(draw, text: str, max_w: int, max_h: int, start_size: int, min_size: int, bold=False, italic=False):
    """Auto-shrink the font size until text fits the given bounding box.
    Uses anchor='lt' so the returned width/height match the actual ink (matches
    how we later draw the text with anchor='lt')."""
    from PIL import ImageFont  # noqa: F401
    size = start_size
    while size >= min_size:
        font = _truetype(size, bold=bold, italic=italic)
        bbox = draw.textbbox((0, 0), text, font=font, anchor="lt")
        w = bbox[2] - bbox[0]
        h = bbox[3] - bbox[1]
        if w <= max_w and h <= max_h:
            return font, w, h
        size -= 4
    font = _truetype(min_size, bold=bold, italic=italic)
    bbox = draw.textbbox((0, 0), text, font=font, anchor="lt")
    return font, bbox[2] - bbox[0], bbox[3] - bbox[1]

@api.get("/exhibitors/me/social-post.png")
async def exhibitor_social_post(user: dict = Depends(require_exhibitor)):
    ex = await db.exhibitors.find_one({"id": user["sub"]}, {"_id": 0, "password_hash": 0})
    if not ex:
        raise HTTPException(status_code=404, detail="Exhibitor not found")

    from PIL import Image, ImageDraw
    template = _get_template()
    if template is None:
        raise HTTPException(status_code=500, detail="Social template missing on server")

    # ---- Cache key (exhibitor + framing + photo mtime) ----
    photo_url = ex.get("profile_photo_url") or ""
    photo_path: Optional[Path] = None
    if photo_url.startswith("/uploads/") or photo_url.startswith("/api/uploads/"):
        photo_path = UPLOAD_DIR / Path(photo_url).name
    photo_mtime = photo_path.stat().st_mtime if (photo_path and photo_path.exists()) else 0
    fx = max(0.0, min(1.0, float(ex.get("photo_focus_x") or 0.5)))
    fy = max(0.0, min(1.0, float(ex.get("photo_focus_y") or 0.35)))
    zoom = max(1.0, min(3.0, float(ex.get("photo_zoom") or 1.0)))
    cache_key = "|".join([
        ex.get("id", ""),
        (ex.get("member_name") or "").strip(),
        (ex.get("business_name") or "").strip(),
        (ex.get("category") or "").strip(),
        (ex.get("position") or "").strip(),
        photo_url, f"{photo_mtime:.0f}",
        f"{fx:.3f}", f"{fy:.3f}", f"{zoom:.3f}",
        f"{_TEMPLATE_CACHE['mtime']:.0f}",
    ])
    cached = _post_cache_get(cache_key)
    if cached:
        return StreamingResponse(io.BytesIO(cached), media_type="image/png", headers={
            "Content-Disposition": f'inline; filename="rama-bazaar-{ex.get("mobile","")}.png"',
            "Cache-Control": "private, max-age=600",
        })

    canvas = Image.new("RGBA", (RENDER_W, RENDER_H), (248, 247, 244, 255))

    # Place profile photo into the silhouette area
    if photo_path and photo_path.exists():
        try:
            photo = Image.open(photo_path).convert("RGB")
            # Pre-shrink very large source photos to bound CPU work (max ~1800px)
            if max(photo.size) > 1800:
                photo.thumbnail((1800, 1800), Image.LANCZOS)
            x0, y0, x1, y1 = SILHOUETTE_BBOX
            target_w, target_h = x1 - x0, y1 - y0
            photo_ratio = photo.width / photo.height
            target_ratio = target_w / target_h
            if photo_ratio > target_ratio:
                scaled_h = int(target_h * zoom)
                scaled_w = int(scaled_h * photo_ratio)
            else:
                scaled_w = int(target_w * zoom)
                scaled_h = int(scaled_w / photo_ratio)
            scaled = photo.resize((scaled_w, scaled_h), Image.LANCZOS)
            ox = int(fx * (scaled_w - target_w))
            oy = int(fy * (scaled_h - target_h))
            ox = max(0, min(scaled_w - target_w, ox))
            oy = max(0, min(scaled_h - target_h, oy))
            crop = scaled.crop((ox, oy, ox + target_w, oy + target_h))
            canvas.paste(crop, (x0, y0))
        except Exception as e:
            logger.warning(f"social-post photo paste failed: {e}")

    # Composite the template over the photo (template transparency reveals photo)
    canvas.alpha_composite(template)

    # Draw participation text inside the rectangle
    draw = ImageDraw.Draw(canvas)
    tx0, ty0, tx1, ty1 = TEXT_BOX
    box_w = tx1 - tx0
    box_h = ty1 - ty0

    name = (ex.get("member_name") or "").strip().upper() or "MEMBER"
    position = (ex.get("position") or "").strip()
    company = (ex.get("business_name") or "").strip()
    category = (ex.get("category") or "").strip().upper()

    # Vertical budgets (relative — works at any RENDER_SCALE)
    name_h_target = int(box_h * 0.34)
    pos_h_target = int(box_h * 0.13)
    co_h_target = int(box_h * 0.22)
    cat_h_target = int(box_h * 0.12)

    s = RENDER_SCALE
    name_font, name_w, name_h = _fit_text(draw, name, box_w - int(80*s), name_h_target, start_size=int(210*s), min_size=int(120*s), bold=True)
    pos_font, pos_w, pos_h = _fit_text(draw, position or " ", box_w - int(100*s), pos_h_target, start_size=int(72*s), min_size=int(48*s), italic=True)
    co_font, co_w, co_h = _fit_text(draw, company, box_w - int(80*s), co_h_target, start_size=int(130*s), min_size=int(78*s), bold=True)
    # Category uses Cinzel
    cat_font_size = int(70 * s)
    cat_min = int(40 * s)
    while cat_font_size >= cat_min:
        cf = _cinzel(cat_font_size)
        cbb = draw.textbbox((0, 0), category, font=cf, anchor="lt")
        if (cbb[2] - cbb[0]) <= box_w - int(120*s) and (cbb[3] - cbb[1]) <= cat_h_target:
            break
        cat_font_size -= max(2, int(4 * s))
    cat_font = _cinzel(cat_font_size)
    cbb = draw.textbbox((0, 0), category, font=cat_font, anchor="lt")
    cat_w, cat_h = cbb[2] - cbb[0], cbb[3] - cbb[1]

    gap = max(int(12*s), int(box_h * 0.035))
    total = name_h + (gap + pos_h if position else 0) + gap + co_h + gap + cat_h
    cy = ty0 + max(0, (box_h - total) // 2)

    draw.text((tx0 + (box_w - name_w) // 2, cy), name, font=name_font, fill=NAVY_INK, anchor="lt")
    cy += name_h + gap
    if position:
        draw.text((tx0 + (box_w - pos_w) // 2, cy), position, font=pos_font, fill=GOLD_INK, anchor="lt")
        cy += pos_h + gap
    draw.text((tx0 + (box_w - co_w) // 2, cy), company, font=co_font, fill=NAVY_INK, anchor="lt")
    cy += co_h + gap
    draw.text((tx0 + (box_w - cat_w) // 2, cy), category, font=cat_font, fill=GOLD_INK, anchor="lt")

    buf = io.BytesIO()
    # Balance speed vs file size: compress_level=6 keeps render fast (<1.5s) but produces
    # ~500KB PNGs instead of ~3MB. optimize=True is intentionally OFF (it's 10-20s slower).
    canvas.convert("RGB").save(buf, format="PNG", optimize=False, compress_level=6)
    png_bytes = buf.getvalue()
    _post_cache_set(cache_key, png_bytes)

    fname = f"rama-bazaar-participation-{ex.get('mobile','')}.png"
    return StreamingResponse(io.BytesIO(png_bytes), media_type="image/png", headers={
        "Content-Disposition": f'inline; filename="{fname}"',
        "Cache-Control": "private, max-age=600",
    })

# ---------- Public Roster ----------
@api.get("/roster")
async def roster(category: Optional[str] = None, q: Optional[str] = None):
    filt = {"approved": True, "hidden": {"$ne": True}}
    if category and category != "all":
        filt["category"] = category
    if q:
        filt["$or"] = [
            {"business_name": {"$regex": q, "$options": "i"}},
            {"member_name": {"$regex": q, "$options": "i"}},
            {"category": {"$regex": q, "$options": "i"}},
            {"description": {"$regex": q, "$options": "i"}},
        ]
    exhibitors = await db.exhibitors.find(filt, {"_id": 0, "password_hash": 0}).sort([("featured", -1), ("created_at", 1)]).to_list(1000)
    exhibitors = [public_exhibitor(e) for e in exhibitors]
    ads = await db.sponsor_ads.find({"placement": "inline", "active": True}, {"_id": 0}).sort("order", 1).to_list(50)
    return {"exhibitors": exhibitors, "inline_ads": ads}

@api.get("/roster/categories")
async def roster_categories():
    cats = await db.exhibitors.distinct("category", {"approved": True, "hidden": {"$ne": True}})
    return sorted([c for c in cats if c])

@api.get("/roster/sponsors")
async def public_sponsors():
    popup = await db.sponsor_ads.find({"placement": "popup", "active": True}, {"_id": 0}).sort("order", 1).to_list(10)
    footer = await db.sponsor_ads.find({"placement": "footer", "active": True}, {"_id": 0}).sort("order", 1).to_list(10)
    featured = await db.sponsor_ads.find({"placement": "featured", "active": True}, {"_id": 0}).sort("order", 1).to_list(10)
    return {"popup": popup, "footer": footer, "featured": featured}

@api.post("/sponsor-ads/{ad_id}/impression")
async def sponsor_impression(ad_id: str):
    await db.sponsor_ads.update_one({"id": ad_id}, {"$inc": {"impressions": 1}})
    return {"ok": True}

@api.post("/sponsor-ads/{ad_id}/click")
async def sponsor_click(ad_id: str):
    await db.sponsor_ads.update_one({"id": ad_id}, {"$inc": {"clicks": 1}})
    return {"ok": True}

# ---------- Admin: Exhibitors ----------
@api.get("/admin/exhibitors")
async def admin_list_exhibitors(_: dict = Depends(require_admin)):
    items = await db.exhibitors.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", -1).to_list(2000)
    return items

@api.put("/admin/exhibitors/{ex_id}")
async def admin_update_exhibitor(ex_id: str, payload: dict, _: dict = Depends(require_admin)):
    payload.pop("id", None)
    payload.pop("password_hash", None)
    payload["updated_at"] = now_iso()
    await db.exhibitors.update_one({"id": ex_id}, {"$set": payload})
    ex = await db.exhibitors.find_one({"id": ex_id}, {"_id": 0, "password_hash": 0})
    return ex

@api.delete("/admin/exhibitors/{ex_id}")
async def admin_delete_exhibitor(ex_id: str, _: dict = Depends(require_admin)):
    await db.exhibitors.delete_one({"id": ex_id})
    return {"ok": True}

@api.post("/admin/exhibitors/{ex_id}/reset-password")
async def admin_reset_pw(ex_id: str, data: PasswordResetIn, _: dict = Depends(require_admin)):
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="Password too short")
    await db.exhibitors.update_one({"id": ex_id}, {"$set": {"password_hash": hash_password(data.new_password)}})
    return {"ok": True}

# ---------- Admin: Visitors ----------
@api.get("/admin/visitors")
async def admin_list_visitors(_: dict = Depends(require_admin)):
    items = await db.visitors.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    return items

@api.get("/admin/visitors/export.csv")
async def admin_export_visitors(_: dict = Depends(require_admin)):
    items = await db.visitors.find({}, {"_id": 0}).sort("created_at", -1).to_list(50000)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Full Name", "Mobile", "Business", "Industry", "City", "Referred By", "Email", "LVB Member", "LVB Chapter", "Photo URL", "Attended", "Attended At", "Created At"])
    for v in items:
        writer.writerow([v.get("full_name"), v.get("mobile"), v.get("business_name"), v.get("industry"), v.get("city"), v.get("referred_by"), v.get("email"), "Yes" if v.get("is_lvb_member") else "No", v.get("lvb_chapter", ""), v.get("photo_url", ""), v.get("attended"), v.get("attended_at"), v.get("created_at")])
    return StreamingResponse(io.BytesIO(buf.getvalue().encode("utf-8")), media_type="text/csv", headers={"Content-Disposition": 'attachment; filename="visitors.csv"'})

@api.delete("/admin/visitors/{vid}")
async def admin_delete_visitor(vid: str, _: dict = Depends(require_admin)):
    await db.visitors.delete_one({"id": vid})
    return {"ok": True}

# ---------- Admin: Exhibitors Bundle Export (logos + photos + Excel sheet) ----------
@api.get("/admin/exhibitors/export.zip")
async def admin_export_exhibitors_bundle(request: Request, base: Optional[str] = None, _: dict = Depends(require_admin)):
    import zipfile
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    items = await db.exhibitors.find({}, {"_id": 0, "password_hash": 0}).sort("business_name", 1).to_list(50000)

    # URL precedence: explicit ?base= → PUBLIC_BASE_URL env → Origin/X-Forwarded-Host → request.url
    base = (
        (base or "").rstrip("/")
        or os.environ.get("PUBLIC_BASE_URL", "").rstrip("/")
        or (request.headers.get("origin") or "").rstrip("/")
        or (f"https://{request.headers.get('x-forwarded-host')}" if request.headers.get("x-forwarded-host") else "")
        or f"{request.url.scheme}://{request.url.netloc}"
    )

    def _resolve_upload_path(url: str):
        """Map an /api/uploads/... or /uploads/... URL to a local file path under UPLOAD_DIR."""
        if not url:
            return None
        # strip query string + leading slashes
        u = url.split("?", 1)[0]
        for prefix in ("/api/uploads/", "/uploads/"):
            if u.startswith(prefix):
                fname = u[len(prefix):]
                p = UPLOAD_DIR / fname
                if p.exists() and p.is_file():
                    return p
        return None

    def _safe(name: str) -> str:
        return "".join(c if (c.isalnum() or c in "-_") else "_" for c in (name or "").strip())[:60] or "exhibitor"

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Exhibitors"
    headers = [
        "Business Name", "Member Name", "Position", "Category", "Mobile", "WhatsApp",
        "Email", "Website", "Instagram", "Facebook", "LinkedIn",
        "Slug", "Digital Card URL", "Approved", "Paid",
        "Logo File", "Profile Photo File", "Description",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1B194B")
    header_font = Font(color="FBF6E8", bold=True, size=11)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    widths = [28, 22, 18, 18, 14, 14, 26, 28, 22, 22, 22, 10, 56, 10, 8, 28, 28, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else ("A" + chr(64 + i - 26))].width = w

    # Will write ZIP to memory
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for ex in items:
            slug = ex.get("slug") or ""
            biz = _safe(ex.get("business_name") or ex.get("member_name") or slug or "exhibitor")
            digi_url = f"{base}/c/{slug}" if slug else ""

            # Copy logo
            logo_filename = ""
            lp = _resolve_upload_path(ex.get("logo_url") or "")
            if lp:
                ext = lp.suffix.lower().lstrip(".")
                logo_filename = f"logos/{biz}-{slug}-logo.{ext}"
                try:
                    zf.write(str(lp), logo_filename)
                except Exception:
                    logo_filename = ""

            # Copy profile photo
            photo_filename = ""
            pp = _resolve_upload_path(ex.get("profile_photo_url") or "")
            if pp:
                ext = pp.suffix.lower().lstrip(".")
                photo_filename = f"photos/{biz}-{slug}-photo.{ext}"
                try:
                    zf.write(str(pp), photo_filename)
                except Exception:
                    photo_filename = ""

            ws.append([
                ex.get("business_name") or "",
                ex.get("member_name") or "",
                ex.get("position") or "",
                ex.get("category") or "",
                ex.get("mobile") or "",
                ex.get("whatsapp") or "",
                ex.get("email") or "",
                ex.get("website") or "",
                ex.get("instagram") or "",
                ex.get("facebook") or "",
                ex.get("linkedin") or "",
                slug,
                digi_url,
                "Yes" if ex.get("approved") else "No",
                "Yes" if ex.get("paid") else "No",
                logo_filename,
                photo_filename,
                ex.get("description") or "",
            ])

        # Wrap text in description col, freeze header
        ws.freeze_panes = "A2"
        last_col_letter = "R"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=(cell.column_letter == last_col_letter))

        # Save xlsx into the ZIP
        xbuf = io.BytesIO()
        wb.save(xbuf)
        xbuf.seek(0)
        zf.writestr("exhibitors.xlsx", xbuf.getvalue())

        # Small README
        readme = (
            "Rama Bazaar 1.0 — Exhibitors Bundle\n"
            f"Generated: {now_iso()}\n"
            f"Total exhibitors: {len(items)}\n\n"
            "Contents:\n"
            "  /exhibitors.xlsx   — Master sheet (all fields + digital card URL per exhibitor)\n"
            "  /logos/            — Each exhibitor's logo (named: <business>-<slug>-logo.<ext>)\n"
            "  /photos/           — Each exhibitor's profile photo (named: <business>-<slug>-photo.<ext>)\n"
            "\nDigital card URLs follow the pattern: " + base + "/c/<slug>\n"
        )
        zf.writestr("README.txt", readme)

    zbuf.seek(0)
    return StreamingResponse(
        zbuf,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="rama-bazaar-exhibitors.zip"'},
    )


# ---------- Admin: Exhibitor Badges (3.5" × 5", premium PNG, bundled as ZIP) ----------
def _render_exhibitor_badge(ex: dict) -> bytes:
    """Render a single 3.5" × 5" portrait PNG badge at 300 DPI (1050 × 1500 px)."""
    from PIL import Image, ImageDraw
    W, H = 1050, 1500
    bg = Image.new("RGB", (W, H), "#f5efe1")
    draw = ImageDraw.Draw(bg, "RGBA")

    brand = ROOT_DIR / "assets" / "brand"

    # ------- helpers -------
    def _paste(path, max_w, max_h, cx, cy, auto_crop=False):
        try:
            im = Image.open(str(path)).convert("RGBA")
        except Exception:
            return None
        if auto_crop:
            bb = im.getbbox()
            if bb:
                im = im.crop(bb)
        iw, ih = im.size
        s = min(max_w / iw, max_h / ih)
        new = (max(1, int(iw * s)), max(1, int(ih * s)))
        im = im.resize(new, Image.LANCZOS)
        bg.paste(im, (cx - new[0] // 2, cy - new[1] // 2), im)
        return new

    # ------- Outer foil frame (single subtle layered border) -------
    draw.rounded_rectangle([(22, 22), (W - 22, H - 22)], radius=24,
                           outline="#d8bc84", width=2)
    draw.rounded_rectangle([(32, 32), (W - 32, H - 32)], radius=18,
                           outline="#e7d2a6", width=1)

    # ============ ZONE A · Sponsor strip ============
    draw.text((W // 2, 78), "POWERED BY",
              font=_cinzel(24), fill="#b2873d", anchor="mt")
    _paste(brand / "coco-salons.jpg", max_w=440, max_h=130, cx=W // 2, cy=178)
    # Single hairline divider closes the sponsor zone
    draw.line([(180, 258), (W - 180, 258)], fill="#d8bc84", width=1)

    # ============ ZONE B · Hero brand lockup ============
    # Auto-crop the lockup's transparent padding so the visual reads tight
    _paste(brand / "rama-bazaar-lockup.png", max_w=520, max_h=400,
           cx=W // 2, cy=500, auto_crop=True)

    # Delicate ornament between the hero and the crest
    orn_y = 735
    draw.line([(W // 2 - 140, orn_y), (W // 2 - 22, orn_y)],
              fill="#d8bc84", width=1)
    draw.line([(W // 2 + 22, orn_y), (W // 2 + 140, orn_y)],
              fill="#d8bc84", width=1)
    pts = [(W // 2, orn_y - 6), (W // 2 + 6, orn_y),
           (W // 2, orn_y + 6), (W // 2 - 6, orn_y)]
    draw.polygon(pts, fill="#b2873d")

    # ============ ZONE C · Exhibitor crest ============
    crest_size = 235
    crest_x = (W - crest_size) // 2
    crest_y = 760
    draw.rounded_rectangle(
        [(crest_x, crest_y), (crest_x + crest_size, crest_y + crest_size)],
        radius=26, fill="#fbf8f0", outline="#d8bc84", width=2,
    )

    logo_path = None
    logo_url = ex.get("logo_url") or ""
    if logo_url:
        u = logo_url.split("?", 1)[0]
        for pref in ("/api/uploads/", "/uploads/"):
            if u.startswith(pref):
                p = UPLOAD_DIR / u[len(pref):]
                if p.exists():
                    logo_path = p
                    break
    if logo_path:
        _paste(logo_path, max_w=crest_size - 64, max_h=crest_size - 64,
               cx=crest_x + crest_size // 2, cy=crest_y + crest_size // 2)
    else:
        initial_f = _truetype(180, italic=True)
        initial = (ex.get("business_name") or ex.get("member_name") or "R").strip()[:1].upper()
        draw.text((crest_x + crest_size // 2, crest_y + crest_size // 2 + 8),
                  initial, font=initial_f, fill="#b2873d", anchor="mm")

    # ============ ZONE D · Exhibitor identity ============
    text_top = crest_y + crest_size + 44  # ≈ 1036

    # Name — italic Playfair, navy
    member = (ex.get("member_name") or "").strip() or "Exhibitor"
    name_f, _nw, name_h = _fit_text(draw, member[:34], max_w=W - 220, max_h=90,
                                    start_size=72, min_size=44, italic=True)
    draw.text((W // 2, text_top), member[:34], font=name_f,
              fill="#1B194B", anchor="mt")
    next_y = text_top + name_h + 14

    # Position — tracked caps, muted
    position = (ex.get("position") or "").strip()
    if position:
        pos_f = _cinzel(20)
        draw.text((W // 2, next_y), position[:42].upper(),
                  font=pos_f, fill="#7a7868", anchor="mt")
        next_y += 34

    # Business name — Cinzel caps in gold, auto-fit width
    biz = (ex.get("business_name") or "").strip()
    if biz:
        biz_size = 38
        while biz_size >= 24:
            f = _cinzel(biz_size)
            tb = draw.textbbox((0, 0), biz[:36], font=f, anchor="lt")
            if (tb[2] - tb[0]) <= (W - 220):
                break
            biz_size -= 3
        biz_font = _cinzel(biz_size)
        draw.text((W // 2, next_y + 6), biz[:36],
                  font=biz_font, fill="#b2873d", anchor="mt")
        bb = draw.textbbox((0, 0), biz[:36], font=biz_font, anchor="lt")
        next_y += (bb[3] - bb[1]) + 24

    # Phone pill — luxury chip with tracked digits
    phone = (ex.get("whatsapp") or ex.get("mobile") or "").strip()
    if phone:
        phone_disp = f"+91 {phone[:5]} {phone[5:10]}" if phone.isdigit() and len(phone) == 10 else f"+91 {phone}"
        ph_f = _truetype(32)
        tb = draw.textbbox((0, 0), phone_disp, font=ph_f, anchor="lt")
        ph_w = tb[2] - tb[0]
        pad_x, pad_y = 28, 12
        pill_w = ph_w + pad_x * 2
        pill_h = (tb[3] - tb[1]) + pad_y * 2
        pill_y = next_y + 4
        pill_x0 = (W - pill_w) // 2
        draw.rounded_rectangle(
            [(pill_x0, pill_y), (pill_x0 + pill_w, pill_y + pill_h)],
            radius=pill_h // 2, fill="#fbf8f0", outline="#d8bc84", width=1,
        )
        draw.text((W // 2, pill_y + pill_h // 2),
                  phone_disp, font=ph_f, fill="#1B194B", anchor="mm")

    # ============ ZONE E · Footer endorsement ============
    foot_top = H - 182
    draw.line([(180, foot_top), (W - 180, foot_top)], fill="#d8bc84", width=1)
    _paste(brand / "lvb-rama-ink.png", max_w=180, max_h=52,
           cx=W // 2, cy=foot_top + 46)
    draw.text((W // 2, foot_top + 90), "AN LVB RAMA INITIATIVE",
              font=_cinzel(16), fill="#7a7868", anchor="mt")
    slug = ex.get("slug") or ""
    if slug:
        sub_f = _truetype(16, italic=True)
        draw.text((W // 2, foot_top + 124), f"Ref · {slug.upper()}",
                  font=sub_f, fill="#9a9685", anchor="mt")

    buf = io.BytesIO()
    bg.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


@api.get("/admin/exhibitors/badges.zip")
async def admin_export_exhibitor_badges(_: dict = Depends(require_admin)):
    import zipfile
    items = await db.exhibitors.find({}, {"_id": 0, "password_hash": 0}).sort("business_name", 1).to_list(50000)

    def _safe(name: str) -> str:
        return "".join(c if (c.isalnum() or c in "-_") else "_" for c in (name or "").strip())[:60] or "exhibitor"

    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for ex in items:
            png = _render_exhibitor_badge(ex)
            slug = ex.get("slug") or ""
            biz = _safe(ex.get("business_name") or ex.get("member_name") or slug or "exhibitor")
            fname = f"{biz}-{slug}-badge.png" if slug else f"{biz}-badge.png"
            zf.writestr(fname, png)
        zf.writestr("README.txt",
                    "Rama Bazaar 1.0 — Exhibitor Badges\n"
                    f"Generated: {now_iso()}\n"
                    f"Total badges: {len(items)}\n\n"
                    "Each PNG is 1050 × 1500 px (3.5\" × 5\" at 300 DPI), portrait orientation,\n"
                    "ready for direct printing on standard name-tag stock.\n")
    zbuf.seek(0)
    return StreamingResponse(
        zbuf,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="rama-bazaar-exhibitor-badges.zip"'},
    )


# ---------- Admin: Sponsor Ads ----------
@api.get("/admin/sponsor-ads")
async def admin_list_ads(_: dict = Depends(require_admin)):
    items = await db.sponsor_ads.find({}, {"_id": 0}).sort("order", 1).to_list(200)
    return items

@api.post("/admin/sponsor-ads")
async def admin_create_ad(data: SponsorAdIn, _: dict = Depends(require_admin)):
    doc = data.model_dump()
    doc["id"] = uuid.uuid4().hex
    doc["impressions"] = 0
    doc["clicks"] = 0
    doc["created_at"] = now_iso()
    await db.sponsor_ads.insert_one(dict(doc))
    return doc

@api.put("/admin/sponsor-ads/{ad_id}")
async def admin_update_ad(ad_id: str, data: SponsorAdIn, _: dict = Depends(require_admin)):
    upd = data.model_dump()
    upd["updated_at"] = now_iso()
    await db.sponsor_ads.update_one({"id": ad_id}, {"$set": upd})
    ad = await db.sponsor_ads.find_one({"id": ad_id}, {"_id": 0})
    return ad

@api.delete("/admin/sponsor-ads/{ad_id}")
async def admin_delete_ad(ad_id: str, _: dict = Depends(require_admin)):
    await db.sponsor_ads.delete_one({"id": ad_id})
    return {"ok": True}

# ---------- Admin: Stats ----------
@api.get("/admin/stats")
async def admin_stats(_: dict = Depends(require_admin)):
    total_visitors = await db.visitors.count_documents({})
    present = await db.visitors.count_documents({"attended": True})
    total_ex = await db.exhibitors.count_documents({})
    approved_ex = await db.exhibitors.count_documents({"approved": True})
    paid_ex = await db.exhibitors.count_documents({"paid": True})
    s = await get_settings()
    limit = int(s.get("exhibitor_limit", 100))
    ads = await db.sponsor_ads.find({}, {"_id": 0}).to_list(200)
    total_imp = sum(a.get("impressions", 0) for a in ads)
    total_clicks = sum(a.get("clicks", 0) for a in ads)
    return {
        "total_visitors": total_visitors,
        "present_visitors": present,
        "pending_visitors": total_visitors - present,
        "total_exhibitors": total_ex,
        "approved_exhibitors": approved_ex,
        "paid_exhibitors": paid_ex,
        "unpaid_exhibitors": total_ex - paid_ex,
        "exhibitor_limit": limit,
        "remaining_slots": max(0, limit - total_ex),
        "sponsor_impressions": total_imp,
        "sponsor_clicks": total_clicks,
    }

# ---------- WhatsApp (BizChat — Meta-approved templates) ----------
def _bizchat_config(s: dict):
    vendor = s.get("bizchat_vendor_uid") or os.environ.get("BIZCHAT_VENDOR_UID")
    token = s.get("bizchat_token") or os.environ.get("BIZCHAT_TOKEN")
    base = os.environ.get("BIZCHAT_API_BASE", "https://bizchatapi.in/api")
    return base, vendor, token

async def send_bizchat_template(to_mobile: str, template_name: str, header_image: Optional[str], fields: List[str], name: str = "", template_language: Optional[str] = None):
    """Send a Meta-approved template message via BizChat.

    fields are body variables in order ({{1}}, {{2}} ...). If the template has a media
    header (image), pass it via header_image. template_language overrides the global
    setting when provided (e.g. "en", "en_US").
    """
    s = await get_settings()
    base, vendor, token = _bizchat_config(s)
    if not vendor or not token or not template_name:
        logger.info("BizChat template send skipped (missing vendor / token / template)")
        return {"skipped": True, "reason": "bizchat-not-configured-or-template-missing"}
    lang = (template_language or s.get("bizchat_template_language") or "en").strip()
    from_id = s.get("bizchat_from_phone_id") or ""
    url = f"{base}/{vendor}/contact/send-template-message?token={token}"
    payload = {
        "phone_number": to_mobile,
        "template_name": template_name,
        "template_language": lang,
    }
    if from_id:
        payload["from_phone_number_id"] = from_id
    if header_image:
        payload["header_image"] = header_image
        payload["header_field_1"] = name or ""
    for i, val in enumerate(fields or [], start=1):
        payload[f"field_{i}"] = val
    if name:
        first, _, last = name.partition(" ")
        payload["contact"] = {
            "first_name": first or name,
            "last_name": last,
            "country": "india",
            "language_code": lang,
        }
    try:
        async with httpx.AsyncClient(timeout=20) as c:
            r = await c.post(url, json=payload)
            body_text = r.text[:1500]
            body_json = None
            try:
                body_json = r.json()
            except Exception:
                body_json = None
            # Detect BizChat soft errors (HTTP 200 but provider says fail)
            provider_ok = True
            provider_msg = ""
            if isinstance(body_json, dict):
                # Common shapes:
                #   {status: success|error, message: "..."}
                #   {result: success|failed, message: "..."}   <-- bizchatapi.in
                #   {success: true|false}
                #   {error: "..."}
                st = body_json.get("status")
                if isinstance(st, str) and st.lower() in ("error", "failed", "fail"):
                    provider_ok = False
                res_field = body_json.get("result")
                if isinstance(res_field, str) and res_field.lower() in ("error", "failed", "fail"):
                    provider_ok = False
                if body_json.get("success") is False:
                    provider_ok = False
                if body_json.get("error"):
                    provider_ok = False
                provider_msg = body_json.get("message") or body_json.get("error") or ""
            return {
                "status": r.status_code,
                "provider_ok": provider_ok,
                "provider_msg": provider_msg,
                "body": body_text,
                "body_json": body_json,
                "payload_sent": payload,
            }
    except Exception as e:
        logger.warning(f"BizChat send error: {e}")
        return {"error": str(e), "provider_ok": False}

@api.get("/admin/bizchat/templates")
async def bizchat_templates(_: dict = Depends(require_admin)):
    s = await get_settings()
    base, vendor, token = _bizchat_config(s)
    if not vendor or not token:
        raise HTTPException(status_code=400, detail="BizChat is not configured. Add vendor UID and token in Settings.")
    url = f"{base}/{vendor}/contact/template-list?token={token}"
    try:
        async with httpx.AsyncClient(timeout=20) as c:
            r = await c.get(url)
            try:
                data = r.json()
            except Exception:
                data = {"raw": r.text}
            return {"status": r.status_code, "data": data}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"BizChat error: {e}")

@api.post("/admin/bizchat/test-send")
async def bizchat_test_send(payload: dict, _: dict = Depends(require_admin)):
    """Quick template test from admin. Payload: { mobile, template?, name? }"""
    mobile = normalize_mobile(payload.get("mobile", ""))
    if len(mobile) != 10:
        raise HTTPException(status_code=400, detail="Invalid mobile number")
    s = await get_settings()
    template = payload.get("template") or s.get("bizchat_template_visitor") or ""
    name = payload.get("name", "Friend")
    to = "91" + mobile
    res = await send_bizchat_template(to, template, header_image=None, fields=[name], name=name)
    return {"sent": True, "result": res}

@api.post("/visitors/send-whatsapp/{qr_id}")
async def send_whatsapp(qr_id: str, request: Request):
    v = await db.visitors.find_one({"qr_id": qr_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Visitor not found")
    s = await get_settings()
    template = s.get("bizchat_template_visitor", "")
    if not template:
        return {"sent": False, "result": {"skipped": True, "reason": "visitor template not set in Admin → Settings → BizChat"}}
    base_url = str(request.base_url).rstrip("/")
    qr_url = f"{base_url}/api/visitors/qr/{qr_id}.png"
    to = ("91" + v["mobile"]) if len(v["mobile"]) == 10 else v["mobile"]
    result = await send_bizchat_template(
        to_mobile=to,
        template_name=template,
        header_image=qr_url,           # template's header IMAGE variable receives the branded QR poster
        fields=[v.get("full_name", "")],  # {{1}} in body = name (override in your approved template as needed)
        name=v.get("full_name", ""),
    )
    return {"sent": True, "result": result}


# ---------- BizChat Broadcast (bulk send with personalised event pass) ----------
def _public_base_url(request: Request) -> str:
    """Return the externally reachable base URL (uses REACT_APP_BACKEND_URL when set)."""
    env_base = os.environ.get("PUBLIC_BASE_URL") or os.environ.get("REACT_APP_BACKEND_URL")
    if env_base:
        return env_base.rstrip("/")
    return str(request.base_url).rstrip("/")


def _interpolate(template_str: str, ctx: dict) -> str:
    """Replace {name}, {business_name}, {city}, {mobile}, {industry}, {category} tokens."""
    if not template_str:
        return ""
    out = template_str
    for k, v in ctx.items():
        out = out.replace("{" + k + "}", str(v or ""))
    return out


async def _broadcast_audience(audience: str) -> List[dict]:
    """Resolve the audience into a list of recipient dicts with {mobile, name, ctx, pass_url_key}."""
    out: List[dict] = []
    if audience.startswith("visitors_"):
        q = {}
        if audience == "visitors_present":
            q = {"attended": True}
        elif audience == "visitors_pending":
            q = {"attended": {"$ne": True}}
        # else: visitors_all
        cur = db.visitors.find(q, {"_id": 0})
        async for v in cur:
            mob = (v.get("mobile") or "").strip()
            if not mob:
                continue
            out.append({
                "kind": "visitor",
                "mobile": mob,
                "name": v.get("full_name") or "",
                "qr_id": v.get("qr_id") or "",
                "ctx": {
                    "name": v.get("full_name") or "",
                    "business_name": v.get("business_name") or "",
                    "city": v.get("city") or "",
                    "industry": v.get("industry") or "",
                    "mobile": mob,
                },
            })
    elif audience.startswith("exhibitors_"):
        q = {"hidden": {"$ne": True}}
        if audience == "exhibitors_paid":
            q["paid"] = True
        elif audience == "exhibitors_approved":
            q["approved"] = True
        cur = db.exhibitors.find(q, {"_id": 0, "password_hash": 0})
        async for ex in cur:
            mob = (ex.get("whatsapp") or ex.get("mobile") or "").strip()
            if not mob:
                continue
            out.append({
                "kind": "exhibitor",
                "mobile": mob,
                "name": ex.get("member_name") or "",
                "qr_id": "",
                "ctx": {
                    "name": ex.get("member_name") or "",
                    "business_name": ex.get("business_name") or "",
                    "city": ex.get("city") or "",
                    "industry": ex.get("category") or "",
                    "category": ex.get("category") or "",
                    "mobile": mob,
                },
            })
    return out


@api.get("/admin/bizchat/audience-count")
async def bizchat_audience_count(audience: str, _: dict = Depends(require_admin)):
    """Return how many recipients a given audience selector resolves to."""
    recipients = await _broadcast_audience(audience)
    return {"audience": audience, "count": len(recipients)}


@api.post("/admin/bizchat/broadcast")
async def bizchat_broadcast(payload: dict, request: Request, _: dict = Depends(require_admin)):
    """Send a Meta-approved template to a chosen audience with a personalised header image.

    Payload:
      template_name (str)        — Meta-approved template
      audience (str)             — visitors_all | visitors_present | visitors_pending |
                                   exhibitors_all | exhibitors_paid | exhibitors_approved
      field_1..field_5 (str)     — template body variables, may include {name}, {business_name}, {city}, {mobile}
      image_mode (str)           — "personalised_pass" (per visitor) | "shared_url" | "none"
      shared_image_url (str)     — used when image_mode == shared_url
      test_mobiles (list[str])   — if provided, only send to these numbers (subset of audience or override)
      dry_run (bool)             — if true, return the recipient list without sending
    """
    s = await get_settings()
    base, vendor, token = _bizchat_config(s)
    if not vendor or not token:
        raise HTTPException(status_code=400, detail="BizChat not configured. Add Vendor UID + Token in Settings.")

    template_name = (payload.get("template_name") or "").strip()
    if not template_name:
        raise HTTPException(status_code=400, detail="template_name is required")

    template_language = (payload.get("template_language") or "").strip() or None
    audience = (payload.get("audience") or "visitors_all").strip()
    image_mode = (payload.get("image_mode") or "personalised_pass").strip()
    shared_image_url = (payload.get("shared_image_url") or "").strip()
    fields_in = [
        (payload.get("field_1") or "").strip(),
        (payload.get("field_2") or "").strip(),
        (payload.get("field_3") or "").strip(),
        (payload.get("field_4") or "").strip(),
        (payload.get("field_5") or "").strip(),
    ]
    test_mobiles = [normalize_mobile(m) for m in (payload.get("test_mobiles") or []) if m]
    dry_run = bool(payload.get("dry_run", False))

    recipients = await _broadcast_audience(audience)
    if test_mobiles:
        recipients = [r for r in recipients if r["mobile"] in test_mobiles]

    if dry_run:
        return {
            "dry_run": True,
            "audience": audience,
            "total": len(recipients),
            "sample": [{"mobile": r["mobile"], "name": r["name"]} for r in recipients[:10]],
        }

    base_url = _public_base_url(request)
    sem = asyncio.Semaphore(5)  # rate limit concurrency

    results = {"total": len(recipients), "sent": 0, "failed": 0, "errors": [], "by_mobile": [], "samples": []}

    async def _send_one(rec: dict):
        async with sem:
            try:
                # personalise body fields
                personal_fields = [_interpolate(f, rec["ctx"]) for f in fields_in]
                # image header
                if image_mode == "personalised_pass" and rec["kind"] == "visitor" and rec["qr_id"]:
                    header_image = f"{base_url}/api/visitors/qr/{rec['qr_id']}.png"
                elif image_mode == "shared_url" and shared_image_url:
                    header_image = shared_image_url
                else:
                    header_image = None
                mobile_full = "91" + rec["mobile"] if len(rec["mobile"]) == 10 else rec["mobile"]
                res = await send_bizchat_template(
                    to_mobile=mobile_full,
                    template_name=template_name,
                    header_image=header_image,
                    fields=personal_fields,
                    name=rec["name"],
                    template_language=template_language,
                )
                http_ok = isinstance(res, dict) and res.get("status") in (200, 201, 202)
                prov_ok = isinstance(res, dict) and res.get("provider_ok", True)
                ok = http_ok and prov_ok
                if ok:
                    results["sent"] += 1
                else:
                    results["failed"] += 1
                    results["errors"].append({
                        "mobile": rec["mobile"],
                        "name": rec["name"],
                        "http_status": res.get("status") if isinstance(res, dict) else None,
                        "provider_msg": res.get("provider_msg") if isinstance(res, dict) else "",
                        "body": (res.get("body") or "")[:600] if isinstance(res, dict) else str(res),
                    })
                # Always capture first 3 raw responses so admin can inspect what BizChat actually said
                if len(results["samples"]) < 3 and isinstance(res, dict):
                    results["samples"].append({
                        "mobile": rec["mobile"],
                        "name": rec["name"],
                        "http_status": res.get("status"),
                        "provider_ok": res.get("provider_ok"),
                        "provider_msg": res.get("provider_msg"),
                        "body_json": res.get("body_json"),
                        "body_text": (res.get("body") or "")[:600],
                        "sent_header_image": (res.get("payload_sent") or {}).get("header_image"),
                    })
                results["by_mobile"].append({
                    "mobile": rec["mobile"], "name": rec["name"], "ok": ok,
                })
            except Exception as e:
                results["failed"] += 1
                results["errors"].append({"mobile": rec.get("mobile"), "error": str(e)})

    await asyncio.gather(*[_send_one(r) for r in recipients])
    # Trim huge response payloads
    results["errors"] = results["errors"][:50]
    results["by_mobile"] = results["by_mobile"][:200]
    return results

# ---------- Committee Members ----------
COMMITTEE_GROUPS = ("rama_bazaar", "management", "supported_by")

class CommitteeMemberIn(BaseModel):
    group: str
    name: str
    role: str = ""
    photo_url: Optional[str] = ""
    logo_url: Optional[str] = ""
    order: Optional[int] = 100
    hidden: Optional[bool] = False

class CommitteeMemberPatch(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    photo_url: Optional[str] = None
    logo_url: Optional[str] = None
    order: Optional[int] = None
    group: Optional[str] = None
    hidden: Optional[bool] = None

def _committee_public(doc: dict) -> dict:
    return {k: doc.get(k) for k in ("id", "group", "name", "role", "photo_url", "logo_url", "order")}

@api.get("/committee")
async def list_committee_public():
    """Public — grouped committee members for the landing page."""
    items = await db.committee.find({"hidden": {"$ne": True}}, {"_id": 0}).sort([("order", 1), ("name", 1)]).to_list(length=200)
    out = {g: [] for g in COMMITTEE_GROUPS}
    for it in items:
        g = it.get("group") or "management"
        if g in out:
            out[g].append(_committee_public(it))
    return out

@api.get("/admin/committee")
async def list_committee_admin(_: dict = Depends(require_admin)):
    items = await db.committee.find({}, {"_id": 0}).sort([("group", 1), ("order", 1), ("name", 1)]).to_list(length=500)
    return items

@api.post("/admin/committee")
async def create_committee_member(data: CommitteeMemberIn, _: dict = Depends(require_admin)):
    if data.group not in COMMITTEE_GROUPS:
        raise HTTPException(status_code=400, detail="Invalid group")
    doc = data.model_dump()
    doc["id"] = uuid.uuid4().hex
    doc["created_at"] = now_iso()
    await db.committee.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/admin/committee/{member_id}")
async def update_committee_member(member_id: str, data: CommitteeMemberPatch, _: dict = Depends(require_admin)):
    payload = {k: v for k, v in data.model_dump().items() if v is not None}
    if "group" in payload and payload["group"] not in COMMITTEE_GROUPS:
        raise HTTPException(status_code=400, detail="Invalid group")
    if not payload:
        raise HTTPException(status_code=400, detail="No fields to update")
    res = await db.committee.update_one({"id": member_id}, {"$set": payload})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Member not found")
    return {"ok": True}

@api.delete("/admin/committee/{member_id}")
async def delete_committee_member(member_id: str, _: dict = Depends(require_admin)):
    res = await db.committee.delete_one({"id": member_id})
    return {"deleted": res.deleted_count}


async def seed_committee():
    """Idempotent seed of the Rama Bazaar 1.0 committee. New rows only — never
    overwrites edits made by admin (photos, role updates, etc.)."""
    seed = [
        # Rama Bazaar Committee
        {"group": "rama_bazaar", "order": 1, "name": "Arpit Jhaveri", "role": "Chair Person (Vice President)"},
        {"group": "rama_bazaar", "order": 2, "name": "Sneha Pankhaniya", "role": "Co Chair Person (LVH)"},
        # Management Committee
        {"group": "management", "order": 1, "name": "Karishma Ghanshani", "role": "Will be entered soon"},
        {"group": "management", "order": 2, "name": "Nishith Shah", "role": "Gifting Partner"},
        {"group": "management", "order": 3, "name": "Harsh Gujarati", "role": "Technology Partner"},
        {"group": "management", "order": 4, "name": "Namrata Revachandani", "role": "Social Media & Marketing Partner"},
        {"group": "management", "order": 5, "name": "Samir Gandhi", "role": "T-Shirt Partner"},
        {"group": "management", "order": 6, "name": "Suresh Prajapati", "role": "Gifting Partner"},
        {"group": "management", "order": 7, "name": "Manish Parekh", "role": "Will be entered soon"},
        {"group": "management", "order": 8, "name": "Neha Chawla", "role": "Will be entered soon"},
        {"group": "management", "order": 9, "name": "Satish Dighe", "role": "Will be entered soon"},
        # Supported By
        {"group": "supported_by", "order": 1, "name": "Khushboo Turkhiya", "role": "President"},
        {"group": "supported_by", "order": 2, "name": "Ashok Chauhan", "role": "Sr."},
    ]
    for s in seed:
        existing = await db.committee.find_one({"name": s["name"], "group": s["group"]})
        if existing:
            continue
        await db.committee.insert_one({
            "id": uuid.uuid4().hex,
            "photo_url": "",
            "logo_url": "",
            "hidden": False,
            "created_at": now_iso(),
            **s,
        })


# ---------- Startup ----------
@app.on_event("startup")
async def startup():
    await db.exhibitors.create_index("mobile", unique=True)
    await db.exhibitors.create_index("slug", unique=True, sparse=True)
    await db.visitors.create_index("mobile", unique=True)
    await db.visitors.create_index("qr_id", unique=True)
    await db.allowed_members.create_index("mobile", unique=True)
    await db.sponsor_ads.create_index("placement")
    # Seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@admin.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin@123")
    existing = await db.admins.find_one({"email": admin_email})
    if not existing:
        await db.admins.insert_one({
            "id": uuid.uuid4().hex,
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "Admin",
            "created_at": now_iso(),
        })
        logger.info(f"Seeded admin: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.admins.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
        logger.info(f"Updated admin password: {admin_email}")
    await get_settings()  # ensure default settings exist
    await seed_committee()

    # One-time migration: rewrite legacy /uploads/* URLs to /api/uploads/* so that
    # the frontend (which routes only /api/* through the ingress) can fetch them.
    for field in ("profile_photo_url", "logo_url", "banner_url"):
        n = await db.exhibitors.update_many(
            {field: {"$regex": "^/uploads/"}},
            [{"$set": {field: {"$concat": ["/api", f"${field}"]}}}],
        )
        if n.modified_count:
            logger.info(f"Migrated {n.modified_count} exhibitor docs: {field} /uploads/ -> /api/uploads/")
    n = await db.sponsor_ads.update_many(
        {"media_url": {"$regex": "^/uploads/"}},
        [{"$set": {"media_url": {"$concat": ["/api", "$media_url"]}}}],
    )
    if n.modified_count:
        logger.info(f"Migrated {n.modified_count} sponsor_ads: media_url /uploads/ -> /api/uploads/")

    # Backfill slugs for any pre-existing exhibitors so each has a short public-card URL
    async for ex in db.exhibitors.find({"$or": [{"slug": {"$exists": False}}, {"slug": ""}, {"slug": None}]}, {"_id": 0, "id": 1}):
        s = await _generate_unique_slug()
        await db.exhibitors.update_one({"id": ex["id"]}, {"$set": {"slug": s}})
        logger.info(f"Assigned slug {s} to exhibitor {ex['id']}")

@app.on_event("shutdown")
async def shutdown():
    client.close()

# ---------- Mount ----------
app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)
